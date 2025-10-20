# -*- coding: utf-8 -*-
"""
Created on Sun Dec 22 01:23:16 2024

@author: 18307
"""
import os
import numpy as np
import pandas as pd

# %% end program
import time
import threading
def shutdown_with_countdown(countdown_seconds=30):
    """
    Initiates a shutdown countdown, allowing the user to cancel shutdown within the given time.

    Args:
        countdown_seconds (int): The number of seconds to wait before shutting down.
    """
    def cancel_shutdown():
        nonlocal shutdown_flag
        user_input = input("\nPress 'c' and Enter to cancel shutdown: ").strip().lower()
        if user_input == 'c':
            shutdown_flag = False
            print("Shutdown cancelled.")

    # Flag to determine whether to proceed with shutdown
    shutdown_flag = True

    # Start a thread to listen for user input
    input_thread = threading.Thread(target=cancel_shutdown, daemon=True)
    input_thread.start()

    # Countdown timer
    print(f"Shutdown scheduled in {countdown_seconds} seconds. Press 'c' to cancel.")
    for i in range(countdown_seconds, 0, -1):
        print(f"Time remaining: {i} seconds", end="\r")
        time.sleep(1)

    # Check the flag after countdown
    if shutdown_flag:
        print("\nShutdown proceeding...")
        os.system("shutdown /s /t 1")  # Execute shutdown command
    else:
        print("\nShutdown aborted.")

def end_program_actions(play_sound=True, shutdown=False, countdown_seconds=120):
    """
    Performs actions at the end of the program, such as playing a sound or shutting down the system.

    Args:
        play_sound (bool): If True, plays a notification sound.
        shutdown (bool): If True, initiates shutdown with a countdown.
        countdown_seconds (int): Countdown time for shutdown confirmation.
    """
    if play_sound:
        try:
            import winsound
            print("Playing notification sound...")
            winsound.Beep(1000, 500)  # Frequency: 1000Hz, Duration: 500ms
        except ImportError:
            print("winsound module not available. Skipping sound playback.")

    if shutdown:
        shutdown_with_countdown(countdown_seconds)

# %% read parameters/save
def read_params(model='exponential', model_fm='basic', model_rcm='differ', folder='fitting_results(15_15_joint_band_from_mat)'):
    identifier = f'{model_fm.lower()}_fm_{model_rcm.lower()}_rcm'
    
    path_current = os.getcwd()
    path_fitting_results = os.path.join(path_current, 'fitting_results', folder)
    file_path = os.path.join(path_fitting_results, f'fitting_results({identifier}).xlsx')
    
    df = pd.read_excel(file_path).set_index('method')
    df_dict = df.to_dict(orient='index')
    
    model = model.upper()
    params = df_dict[model]
    
    return params

from openpyxl import load_workbook
def save_results_to_xlsx_append(results, output_dir, filename, sheet_name='K-Fold Results'):
    """
    Appends results to an existing Excel file or creates a new file if it doesn't exist.

    Args:
        results (list or pd.DataFrame): The results data to save.
        output_dir (str): The directory where the Excel file will be saved.
        filename (str): The name of the Excel file.
        sheet_name (str): The sheet name in the Excel file. Default is 'K-Fold Results'.

    Returns:
        str: The path of the saved Excel file.
    """
    # Convert results to DataFrame if necessary
    if not isinstance(results, pd.DataFrame):
        results_df = pd.DataFrame(results)
    else:
        results_df = results

    # Rearrange columns if "Identifier" is present
    if 'Identifier' in results_df.columns:
        columns_order = ['Identifier'] + [col for col in results_df.columns if col != 'Identifier']
        results_df = results_df[columns_order]

    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)

    # Define the full output path
    output_path = os.path.join(output_dir, filename)

    # Append to existing Excel file or create a new one
    if os.path.exists(output_path):
        print(f"Appending data to existing file: {output_path}")
        with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Get the existing workbook
            existing_workbook = load_workbook(output_path)

            # Check if the sheet exists
            if sheet_name in existing_workbook.sheetnames:
                # Load existing sheet and append
                start_row = existing_workbook[sheet_name].max_row
                results_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=start_row)
            else:
                # Write new sheet if not exists
                results_df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        print(f"Creating new file: {output_path}")
        results_df.to_excel(output_path, index=False, sheet_name=sheet_name)

    print(f"Results successfully saved to: {output_path}")
    return output_path

# %% control; original cm
import torch
import feature_engineering
import cnn_validation
from models import models

def cnn_evaluation_circle_original_cm(feature_cm, subject_range=range(1, 6), experiment_range=range(1, 4), save=False):
    # labels
    labels = utils_feature_loading.read_labels(dataset='seed')
    y = torch.tensor(np.array(labels)).view(-1)
    
    # data and evaluation circle
    all_results_list = []
    
    for sub in subject_range:
        for ex in experiment_range:
            subject_id = f"sub{sub}ex{ex}"
            print(f"Evaluating {subject_id}...")
            
            # CM/MAT
            features = utils_feature_loading.read_fcs_mat('seed', subject_id, feature_cm)
            alpha = features['alpha']
            beta = features['beta']
            gamma = features['gamma']
            
            # CM/H5
            # features = utils_feature_loading.read_fcs('seed', subject_id, feature_cm)
            # alpha = features['alpha']
            # beta = features['beta']
            # gamma = features['gamma']
            
            x = np.stack((alpha, beta, gamma), axis=1)
            
            # cnn model
            cnn_model = models.CNN_2layers_adaptive_maxpool_3()
            # traning and testing
            result_CM = cnn_validation.cnn_cross_validation(cnn_model, x, y)
            
            # Add identifier to the result
            result_flat = {'Identifier': subject_id, **result_CM}
            all_results_list.append(result_flat)
    
    # Convert list of dicts to DataFrame
    df_results = pd.DataFrame(all_results_list)
    
    # Compute mean of all numeric columns (excluding Identifier)
    mean_row = df_results.select_dtypes(include=[np.number]).mean().to_dict()
    mean_row['Identifier'] = 'Average'
    df_results = pd.concat([df_results, pd.DataFrame([mean_row])], ignore_index=True)
    
    # save
    output_dir = os.path.join(os.getcwd(), 'results_cnn_evaluation')
    filename_CM = f"cnn_validation_CM_{feature_cm}.xlsx"
    if save: save_results_to_xlsx_append(df_results, output_dir, filename_CM)
    
    return df_results

# Channel feature
def cnn_evaluation_circle_original_cf(feature, subject_range=range(1, 6), experiment_range=range(1, 4), save=False):
    # labels
    labels = utils_feature_loading.read_labels(dataset='seed')
    y = torch.tensor(np.array(labels)).view(-1)
    
    # data and evaluation circle
    all_results_list = []
    
    for sub in subject_range:
        for ex in experiment_range:
            subject_id = f"sub{sub}ex{ex}"
            print(f"Evaluating {subject_id}...")
            
            # CM/MAT
            features = utils_feature_loading.read_cfs('seed', subject_id, feature)
            alpha = features['alpha']
            beta = features['beta']
            gamma = features['gamma']
            
            # CM/H5
            # features = utils_feature_loading.read_fcs('seed', subject_id, feature_cm)
            # alpha = features['alpha']
            # beta = features['beta']
            # gamma = features['gamma']
            
            x = np.stack((alpha, beta, gamma), axis=1)
            
            # cnn model
            cnn_model = models.CNN_2layers_adaptive_maxpool_3()
            # traning and testing
            result_CM = cnn_validation.cnn_cross_validation(cnn_model, x, y)
            
            # Add identifier to the result
            result_flat = {'Identifier': subject_id, **result_CM}
            all_results_list.append(result_flat)
    
    # Convert list of dicts to DataFrame
    df_results = pd.DataFrame(all_results_list)
    
    # Compute mean of all numeric columns (excluding Identifier)
    mean_row = df_results.select_dtypes(include=[np.number]).mean().to_dict()
    mean_row['Identifier'] = 'Average'
    df_results = pd.concat([df_results, pd.DataFrame([mean_row])], ignore_index=True)
    
    # save
    output_dir = os.path.join(os.getcwd(), 'results_cnn_evaluation')
    filename_CM = f"cnn_validation_CM_{feature}.xlsx"
    if save: save_results_to_xlsx_append(df_results, output_dir, filename_CM)
    
    return df_results

# %% experiment; rebuilded cm
from utils import utils_feature_loading
import spatial_gaussian_smoothing

def cnn_evaluation_circle_rebuilded_cm(feature_cm, projection_params={"source": "auto", "type": "3d"}, sigma=0.05, 
                                       subject_range=range(1, 6), experiment_range=range(1, 4), save=False):
    # labels
    labels = utils_feature_loading.read_labels(dataset='seed')
    y = torch.tensor(np.array(labels)).view(-1)
    
    # data and evaluation circle
    all_results_list = []
    
    for sub in subject_range:
        for ex in experiment_range:
            subject_id = f"sub{sub}ex{ex}"
            print(f"Evaluating {subject_id}...")
            
            # CM/MAT
            features = utils_feature_loading.read_fcs_mat('seed', subject_id, feature_cm)
            alpha = features['alpha']
            beta = features['beta']
            gamma = features['gamma']
            
            # CM/H5
            # features = utils_feature_loading.read_fcs('seed', subject_id, feature_cm)
            # alpha = features['alpha']
            # beta = features['beta']
            # gamma = features['gamma']
            
            alpha_ = spatial_gaussian_smoothing.fcs_spatial_gaussian_smoothing(alpha, projection_params, sigma)
            beta_ = spatial_gaussian_smoothing.fcs_spatial_gaussian_smoothing(beta, projection_params, sigma)
            gamma_ = spatial_gaussian_smoothing.fcs_spatial_gaussian_smoothing(gamma, projection_params, sigma)
            
            alpha = alpha+alpha_
            beta = beta+beta_
            gamma = gamma+gamma_
            
            x = np.stack((alpha, beta, gamma), axis=1)
            
            # cnn model
            cnn_model = models.CNN_2layers_adaptive_maxpool_3()
            # traning and testing
            result_CM = cnn_validation.cnn_cross_validation(cnn_model, x, y)
            
            # Add identifier to the result
            result_flat = {'Identifier': subject_id, **result_CM}
            all_results_list.append(result_flat)
    
    # Convert list of dicts to DataFrame
    df_results = pd.DataFrame(all_results_list)
    
    # Compute mean of all numeric columns (excluding Identifier)
    mean_row = df_results.select_dtypes(include=[np.number]).mean().to_dict()
    mean_row['Identifier'] = 'Average'
    df_results = pd.concat([df_results, pd.DataFrame([mean_row])], ignore_index=True)
    
    # save
    output_dir = os.path.join(os.getcwd(), 'results_cnn_evaluation')
    filename_CM = f"cnn_validation_CM_{feature_cm}_smoothed_{projection_params['source']}_{projection_params['type']}.xlsx"
    if save: save_results_to_xlsx_append(df_results, output_dir, filename_CM)
    
    return df_results

def cnn_evaluation_circle_rebuilded_cf(feature, projection_params={"source": "auto", "type": "3d"}, sigma=0.05, 
                                       subject_range=range(1, 6), experiment_range=range(1, 4), save=False):
    # labels
    labels = utils_feature_loading.read_labels(dataset='seed')
    y = torch.tensor(np.array(labels)).view(-1)
    
    # data and evaluation circle
    all_results_list = []
    
    for sub in subject_range:
        for ex in experiment_range:
            subject_id = f"sub{sub}ex{ex}"
            print(f"Evaluating {subject_id}...")
            
            # CM/MAT
            features = utils_feature_loading.read_cfs('seed', subject_id, feature)
            alpha = features['alpha']
            beta = features['beta']
            gamma = features['gamma']
            
            # CM/H5
            # features = utils_feature_loading.read_fcs('seed', subject_id, feature_cm)
            # alpha = features['alpha']
            # beta = features['beta']
            # gamma = features['gamma']
            
            alpha_ = spatial_gaussian_smoothing.cfs_spatial_gaussian_smoothing(alpha, projection_params, sigma)
            beta_ = spatial_gaussian_smoothing.cfs_spatial_gaussian_smoothing(beta, projection_params, sigma)
            gamma_ = spatial_gaussian_smoothing.cfs_spatial_gaussian_smoothing(gamma, projection_params, sigma)
            
            alpha = alpha+alpha_
            beta = beta+beta_
            gamma = gamma+gamma_
            
            x = np.stack((alpha, beta, gamma), axis=1)
            
            # cnn model
            cnn_model = models.CNN_2layers_adaptive_maxpool_3()
            # traning and testing
            result_CM = cnn_validation.cnn_cross_validation(cnn_model, x, y)
            
            # Add identifier to the result
            result_flat = {'Identifier': subject_id, **result_CM}
            all_results_list.append(result_flat)
    
    # Convert list of dicts to DataFrame
    df_results = pd.DataFrame(all_results_list)
    
    # Compute mean of all numeric columns (excluding Identifier)
    mean_row = df_results.select_dtypes(include=[np.number]).mean().to_dict()
    mean_row['Identifier'] = 'Average'
    df_results = pd.concat([df_results, pd.DataFrame([mean_row])], ignore_index=True)
    
    # save
    output_dir = os.path.join(os.getcwd(), 'results_cnn_evaluation')
    filename_CM = f"cnn_validation_CM_{feature}_smoothed_{projection_params['source']}_{projection_params['type']}.xlsx"
    if save: save_results_to_xlsx_append(df_results, output_dir, filename_CM)
    
    return df_results

if __name__ == '__main__':
    # %% fc
    results_cm = cnn_evaluation_circle_original_cm('pcc', range(1, 6), save=True)
    
    results_cm_smoothed_a2 = cnn_evaluation_circle_rebuilded_cm('pcc', {"source": "manual", "type": "2d"}, 0.05,
                                                             range(1, 6), save=True)
    
    results_cm_smoothed_a3 = cnn_evaluation_circle_rebuilded_cm('pcc', {"source": "auto", "type": "3d"}, 0.05,
                                                             range(1, 6), save=True)
    
    results_cm_smoothed_a2 = cnn_evaluation_circle_rebuilded_cm('pcc', {"source": "auto", "type": "2d"}, 0.05,
                                                             range(1, 6), save=True)
    
    results_cm_smoothed_as = cnn_evaluation_circle_rebuilded_cm('pcc', {"source": "auto", "type": "stereo"}, 0.05,
                                                             range(1, 6), save=True)
    
    # %% cf
    # results_cm = cnn_evaluation_circle_original_cf('de_LDS', range(1, 6), save=True)
    
    # results_cm_smoothed_a2 = cnn_evaluation_circle_rebuilded_cf('de_LDS', {"source": "manual", "type": "2d"}, 0.05,
    #                                                          range(1, 6), save=True)
    
    # results_cm_smoothed_a3 = cnn_evaluation_circle_rebuilded_cf('de_LDS', {"source": "auto", "type": "3d"}, 0.05,
    #                                                          range(1, 6), save=True)
    
    # results_cm_smoothed_a2 = cnn_evaluation_circle_rebuilded_cf('de_LDS', {"source": "auto", "type": "2d"}, 0.05,
    #                                                          range(1, 6), save=True)
    
    # results_cm_smoothed_as = cnn_evaluation_circle_rebuilded_cf('de_LDS', {"source": "auto", "type": "stereo"}, 0.05,
    #                                                          range(1, 6), save=True)
    
    # %% End
    end_program_actions(play_sound=True, shutdown=False, countdown_seconds=120)